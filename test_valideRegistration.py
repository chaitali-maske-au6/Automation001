from selenium.webdriver import Chrome
from Base import InisiteDriver
from Library import ConfigReader
from  Pages import RegistrationPage
import openpyxl
import pytest
from DataGenerate import DataGen
# def dataGenerator():
#     wk=openpyxl.load_workbook("C:/TData10.xlsx")
#     sh=wk['Sheet1']
#     r = sh.max_row
#     li=[]
#     li1=[]
#     for i in range(1,r+1):
#         li1=[]
#         un=sh.cell(i,1)
#         up=sh.cell(i,2)
#         li1.insert(0,un.value)
#         li1.insert(1,up.value)
#         li.insert(i-1,li1)
#     print(li)
#     return li
    # li = [['uname1','pass1'],['uname2','pass2'],['uname3','pass3']]
    # return li
@pytest.mark.parametrze('data',DataGen.dataGenerator())
def test_ValidateRegistration(data):
    driver=InisiteDriver.startBrowser()

    register = RegistrationPage.RegistrationClass(driver)
    register.enter_username(data[0])
    register.enter_password(data[1])



